"""Exportacion a Excel.

PLAN 1.3: si la validacion falla no se entrega un Excel limpio. Las filas
afectadas se marcan y el detalle va en una hoja aparte.
"""

from __future__ import annotations

from collections.abc import Sequence
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill

from contapdf.parsers.auxiliar import Auxiliar
from contapdf.parsers.balanza import Balanza
from contapdf.parsers.estado_cuenta import EstadoCuenta
from contapdf.parsers.mayor import Mayor
from contapdf.parsers.polizas import LibroDiario
from contapdf.validate.rules import NO_VERIFICABLE, Cobertura

_ENCABEZADOS = ("cuenta", "nivel", "cuenta_padre", "naturaleza", "nombre",
                "saldo_ini_deudor", "saldo_ini_acreedor", "debe", "haber",
                "saldo_fin_deudor", "saldo_fin_acreedor", "es_acumulativa")
_MONTOS = ("saldo_ini_deudor", "saldo_ini_acreedor", "debe", "haber",
           "saldo_fin_deudor", "saldo_fin_acreedor")
_FORMATO_MONTO = "#,##0.00"
_ANCHOS = (14, 6, 14, 5, 42, 16, 18, 16, 16, 18, 18, 14)


def _detalle(regla) -> str:
    partes = []
    if regla.exactas:
        partes.append(f"{regla.exactas} exacta"
                      + ("s" if regla.exactas != 1 else ""))
    if regla.con_tolerancia:
        partes.append(f"{len(regla.con_tolerancia)} dentro de tolerancia: "
                      + ", ".join(regla.con_tolerancia[:5]))
    if regla.discrepancias:
        partes.append(f"{len(regla.discrepancias)} con diferencia")
    return "; ".join(partes) or f"{regla.comprobaciones} comprobaciones"


def exportar_balanza(balanza: Balanza, cobertura: Cobertura,
                     destino: Path) -> Path:
    """Escribe el .xlsx en 'destino' y devuelve la ruta.

    La ruta llega como parametro: el nucleo no decide donde escribir, para
    que cada trabajo mande a su propio directorio de tenant.
    """
    libro = Workbook()
    hoja = libro.active
    hoja.title = "Balanza"

    negrita = Font(bold=True)
    alerta = PatternFill(fill_type="solid", start_color="FFF4CCCC",
                         end_color="FFF4CCCC")

    hoja.append(list(_ENCABEZADOS))
    for celda in hoja[1]:
        celda.font = negrita
    for columna, ancho in zip(hoja.iter_cols(min_row=1, max_row=1), _ANCHOS):
        hoja.column_dimensions[columna[0].column_letter].width = ancho
    hoja.freeze_panes = "A2"

    discrepancias = cobertura.discrepancias
    marcadas = {d.indice for d in discrepancias if d.indice >= 0}
    for indice, fila in enumerate(balanza.filas):
        hoja.append([getattr(fila, campo) for campo in _ENCABEZADOS])
        renglon = hoja[hoja.max_row]
        for celda, campo in zip(renglon, _ENCABEZADOS):
            if campo in _MONTOS:
                celda.number_format = _FORMATO_MONTO
            if indice in marcadas:
                celda.fill = alerta

    # La hoja de validacion va siempre, aunque no haya discrepancias: un
    # resultado sin su cobertura no se entrega (PLAN 2).
    detalle = libro.create_sheet("Validacion")
    detalle.append(["regla", "estado", "detalle"])
    for celda in detalle[1]:
        celda.font = negrita
    for regla in cobertura.reglas:
        detalle.append([regla.regla, regla.estado,
                        regla.motivo if regla.estado == NO_VERIFICABLE
                        else _detalle(regla)])
    detalle.append([])
    detalle.append(["fila", "regla", "esperado", "obtenido"])
    for celda in detalle[detalle.max_row]:
        celda.font = negrita
    for d in discrepancias:
        detalle.append([d.fila, d.regla, d.esperado, d.obtenido])
        for celda in detalle[detalle.max_row][2:]:
            celda.number_format = _FORMATO_MONTO
    for columna, ancho in zip(detalle.iter_cols(min_row=1, max_row=1),
                              (18, 18, 60)):
        detalle.column_dimensions[columna[0].column_letter].width = ancho
    detalle.freeze_panes = "A2"

    libro.save(str(destino))
    return destino


_POLIZA = ("poliza_id", "tipo", "naturaleza", "fecha", "descripcion", "folio",
           "total_debe", "total_haber", "completa")
_MOVIMIENTO = ("poliza_id", "orden", "cuenta", "nombre_cuenta", "debe", "haber")
_CFDI = ("poliza_id", "fecha", "documento", "uuid", "rfc", "tipo")
_MONTOS_DIARIO = frozenset({"total_debe", "total_haber", "debe", "haber"})


def _hoja(libro_excel, titulo: str, encabezados, filas, negrita,
          montos=_MONTOS_DIARIO) -> None:
    """Una hoja: encabezado en negrita, montos con formato, panel congelado.

    `filas` puede traer objetos o diccionarios -- la hoja plana se arma como
    diccionario porque mezcla campos de dos tablas. Un campo ausente o en
    None sale VACIO, nunca como cero: un cero inventado es indistinguible de
    uno leido (PLAN 2).
    """
    hoja = libro_excel.create_sheet(titulo)
    hoja.append(list(encabezados))
    for celda in hoja[1]:
        celda.font = negrita
    for fila in filas:
        hoja.append([fila.get(campo) if isinstance(fila, dict)
                     else getattr(fila, campo, None) for campo in encabezados])
        for celda, campo in zip(hoja[hoja.max_row], encabezados):
            if campo in montos:
                celda.number_format = _FORMATO_MONTO
    hoja.freeze_panes = "A2"


def exportar_polizas(libro: LibroDiario, cobertura: Cobertura,
                     destino: Path) -> Path:
    """Tres hojas relacionadas, una plana y la cobertura.

    La plana repite el encabezado de la poliza en cada movimiento: es la
    que el contador filtra. Las otras tres conservan la relacion, que una
    tabla plana pierde.
    """
    libro_excel = Workbook()
    libro_excel.remove(libro_excel.active)
    negrita = Font(bold=True)

    _hoja(libro_excel, "Polizas", _POLIZA, libro.polizas, negrita)
    _hoja(libro_excel, "Movimientos", _MOVIMIENTO, libro.movimientos, negrita)
    _hoja(libro_excel, "CFDI", _CFDI, libro.cfdi, negrita)

    por_id = {p.poliza_id: p for p in libro.polizas}
    planas = []
    for movimiento in libro.movimientos:
        poliza = por_id.get(movimiento.poliza_id)
        fila = {campo: getattr(poliza, campo, None) for campo in _POLIZA}
        fila.update({campo: getattr(movimiento, campo)
                     for campo in _MOVIMIENTO if campo != "poliza_id"})
        planas.append(fila)
    _hoja(libro_excel, "Plana", _POLIZA + _MOVIMIENTO[1:], planas, negrita)

    _validacion(libro_excel, cobertura, negrita)
    libro_excel.save(str(destino))
    return destino


def _validacion(libro_excel, cobertura: Cobertura, negrita) -> None:
    detalle = libro_excel.create_sheet("Validacion")
    detalle.append(["regla", "estado", "detalle"])
    for celda in detalle[1]:
        celda.font = negrita
    for regla in cobertura.reglas:
        detalle.append([regla.regla, regla.estado,
                        regla.motivo if regla.estado == NO_VERIFICABLE
                        else _detalle(regla)])
    detalle.append([])
    detalle.append(["fila", "regla", "esperado", "obtenido"])
    for celda in detalle[detalle.max_row]:
        celda.font = negrita
    for d in cobertura.discrepancias:
        detalle.append([d.fila, d.regla, d.esperado, d.obtenido])
    detalle.freeze_panes = "A2"


_CUENTA_MAYOR = ("cuenta", "nombre_cuenta", "naturaleza", "saldo_inicial",
                 "saldo_final", "total_cargos", "total_abonos")
_MES_MAYOR = ("cuenta", "orden", "periodo", "cargos", "abonos", "saldo",
              "acum_cargos", "acum_abonos")
_MONTOS_MAYOR = frozenset({"saldo_inicial", "saldo_final", "total_cargos",
                           "total_abonos", "cargos", "abonos", "saldo",
                           "acum_cargos", "acum_abonos"})


def exportar_mayor(mayor: Mayor, cobertura: Cobertura, destino: Path) -> Path:
    """Dos hojas relacionadas, una plana y la cobertura."""
    libro = Workbook()
    libro.remove(libro.active)
    negrita = Font(bold=True)

    _hoja(libro, "Cuentas", _CUENTA_MAYOR, mayor.cuentas, negrita,
          _MONTOS_MAYOR)
    _hoja(libro, "Meses", _MES_MAYOR, mayor.meses, negrita, _MONTOS_MAYOR)

    por_cuenta = {c.cuenta: c for c in mayor.cuentas}
    planas = []
    for mes in mayor.meses:
        cuenta = por_cuenta.get(mes.cuenta)
        fila = {campo: getattr(cuenta, campo, None) for campo in _CUENTA_MAYOR}
        fila.update({campo: getattr(mes, campo) for campo in _MES_MAYOR
                     if campo != "cuenta"})
        planas.append(fila)
    _hoja(libro, "Plana", _CUENTA_MAYOR + _MES_MAYOR[1:], planas, negrita,
          _MONTOS_MAYOR)

    _validacion(libro, cobertura, negrita)
    libro.save(str(destino))
    return destino


# La metadata del documento va como columnas de la hoja Cuentas y no en una
# quinta hoja: son de una a tres filas, repetirla no cuesta nada, y una hoja
# de dos renglones que nadie abre es peor que una columna repetida.
_META_EDOCTA = ("banco", "rfc", "periodo_ini", "periodo_fin")
_CUENTA_BANCO = ("num_cuenta", "clabe", "producto", "moneda", "saldo_inicial",
                 "depositos", "retiros", "saldo_corte")
_MOVIMIENTO_BANCO = ("num_cuenta", "dia", "fecha", "descripcion", "referencia",
                     "deposito", "retiro", "saldo", "pagina")
_MONTOS_EDOCTA = frozenset({"saldo_inicial", "depositos", "retiros",
                            "saldo_corte", "deposito", "retiro", "saldo"})


def exportar_estado_cuenta(estado: EstadoCuenta, cobertura: Cobertura,
                           destino: Path) -> Path:
    """Dos hojas relacionadas, una plana y la cobertura.

    Un estado puede traer varias cuentas, asi que la relacion importa: cada
    movimiento apunta a la cuenta que lo contiene y ninguna fila queda
    huerfana. La plana repite el encabezado de la cuenta en cada movimiento,
    que es la que el contador filtra.
    """
    libro = Workbook()
    libro.remove(libro.active)
    negrita = Font(bold=True)

    meta = {campo: getattr(estado.meta, campo) for campo in _META_EDOCTA}
    cuentas = [{**meta, **{campo: getattr(c, campo) for campo in _CUENTA_BANCO}}
               for c in estado.cuentas]
    _hoja(libro, "Cuentas", _META_EDOCTA + _CUENTA_BANCO, cuentas,
          negrita, _MONTOS_EDOCTA)
    _hoja(libro, "Movimientos", _MOVIMIENTO_BANCO, estado.movimientos,
          negrita, _MONTOS_EDOCTA)

    por_cuenta = {c["num_cuenta"]: c for c in cuentas}
    planas = []
    for movimiento in estado.movimientos:
        fila = dict(por_cuenta.get(movimiento.num_cuenta, {}))
        fila.update({campo: getattr(movimiento, campo)
                     for campo in _MOVIMIENTO_BANCO if campo != "num_cuenta"})
        planas.append(fila)
    _hoja(libro, "Plana", _META_EDOCTA + _CUENTA_BANCO
          + _MOVIMIENTO_BANCO[1:], planas, negrita, _MONTOS_EDOCTA)

    _validacion(libro, cobertura, negrita)
    libro.save(str(destino))
    return destino


_FILA_AUXILIAR = ("cuenta", "nombre_cuenta", "saldo_inicial_cuenta", "folio",
                  "fecha", "tipo_movimiento", "documento", "tercero",
                  "concepto", "debe", "haber", "saldo", "es_subtotal",
                  "pagina")
_MONTOS_AUXILIAR = frozenset({"saldo_inicial_cuenta", "debe", "haber", "saldo"})


def exportar_auxiliar(auxiliar: Auxiliar, cobertura: Cobertura,
                      destino: Path) -> Path:
    """Una tabla y la cobertura: el auxiliar ya viene plano.

    La cuenta de la seccion va arrastrada en cada renglon, asi que no hay
    una segunda tabla que relacionar ni una hoja plana que construir.
    """
    libro = Workbook()
    libro.remove(libro.active)
    negrita = Font(bold=True)
    _hoja(libro, "Auxiliar", _FILA_AUXILIAR, auxiliar.filas, negrita,
          _MONTOS_AUXILIAR)
    _validacion(libro, cobertura, negrita)
    libro.save(str(destino))
    return destino

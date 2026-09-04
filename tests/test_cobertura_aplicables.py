"""Cada conteo de cobertura con su denominador.

Fase 7f. `ResultadoRegla` guardaba cuantas comprobaciones corrieron pero no
cuantas PODIA haber corrido, asi que un 5/5 sobre 116 casos y un 116/116 se
imprimian igual. La tabla de resultados de la fase 7d aprobo BBVA con
`saldo_corrido: 5/5` sobre 116 movimientos: la regla corrio en el 4% de la
tabla y se reporto como cuadra.

Es el mismo modo de falla del `734 filas, 0 discrepancias` de balanza-gume,
sobrevivido a la fase 4a. Lo que lo impide ahora no es una convencion sino
dos invariantes con test: `aplicables` no puede faltar en una regla que
cuadra, y ningun conteo se imprime sin su denominador.

Las cifras de este archivo son MEDICIONES tomadas antes de escribir el
campo, calculando el universo fuera del codigo. Si el codigo da otra cosa,
el universo quedo mal; no se ajusta el numero.
"""

from __future__ import annotations

import pytest
from conftest import requires_real_pdf

from contapdf.pipeline import (
    procesar_auxiliar,
    procesar_balanza,
    procesar_estado_cuenta,
    procesar_mayor,
    procesar_polizas,
)
from contapdf.validate.rules import (
    CUADRA,
    NO_VERIFICABLE,
    Cobertura,
    ResultadoRegla,
)


def _regla(cobertura, nombre):
    return next(r for r in cobertura.reglas if r.regla == nombre)


# --- El invariante: cuadrar sin denominador esta prohibido --------------
def test_una_regla_que_cuadra_sin_aplicables_es_ilegal():
    """No es una convencion, es un error de construccion.

    Si una regla no sabe sobre cuantos casos podia correr, no puede
    afirmar que cuadro: es justo lo que hacia el 5/5 de BBVA.
    """
    with pytest.raises(ValueError, match="aplicables"):
        ResultadoRegla(regla="inventada", estado=CUADRA, evaluados=5, exactas=5)


def test_sin_aplicables_si_puede_ser_no_verificable():
    # No saber el universo es una razon legitima para NO afirmar nada.
    regla = ResultadoRegla(regla="inventada", estado=NO_VERIFICABLE,
                           motivo="no se pudo determinar el universo")
    assert regla.aplicables is None


def test_aplicables_nunca_es_menor_que_evaluados():
    with pytest.raises(ValueError, match="aplicables"):
        ResultadoRegla(regla="inventada", estado=CUADRA, aplicables=3,
                       evaluados=5, exactas=5)


# --- comprobaciones: propiedad deprecada, no campo ----------------------
def test_comprobaciones_sigue_leyendose_y_devuelve_evaluados():
    regla = ResultadoRegla(regla="x", estado=CUADRA, aplicables=9, evaluados=7,
                           exactas=7)
    assert regla.comprobaciones == regla.evaluados == 7


def test_comprobaciones_ya_no_se_puede_construir():
    """Dos campos con nombres parecidos y significados distintos es como
    empezo este problema."""
    with pytest.raises(TypeError):
        ResultadoRegla(regla="x", estado=NO_VERIFICABLE, comprobaciones=3)


# --- Ningun conteo sin su denominador -----------------------------------
def test_el_resumen_de_una_regla_lleva_las_dos_cifras():
    regla = ResultadoRegla(regla="saldo_corrido", estado=CUADRA,
                           aplicables=116, evaluados=5, exactas=5)
    texto = regla.resumen()
    assert "5 de 116" in texto
    assert "5 exactos" in texto


def test_el_resumen_de_la_cobertura_lleva_el_denominador():
    cobertura = Cobertura(reglas=(
        ResultadoRegla(regla="saldo_corrido", estado=CUADRA, aplicables=116,
                       evaluados=5, exactas=5),
        ResultadoRegla(regla="resumen", estado=CUADRA, aplicables=1,
                       evaluados=1, exactas=1),
    ))
    texto = cobertura.resumen()
    assert "2 reglas" in texto
    # 6 evaluados de 117 aplicables: el numerador nunca viaja solo.
    assert "6 de 117" in texto


def test_un_universo_sin_determinar_se_dice_en_vez_de_callarse():
    cobertura = Cobertura(reglas=(
        ResultadoRegla(regla="x", estado=NO_VERIFICABLE, motivo="sin universo"),
    ))
    assert "sin determinar" in cobertura.resumen()


# --- Los dos casos que motivaron la fase --------------------------------
def test_bbva_reporta_el_saldo_corrido_sobre_116_no_sobre_5():
    """El caso concreto: el banco imprime el saldo una vez por dia.

    111 de los 116 movimientos no tienen contra que encadenarse. Eso no
    convierte la tabla en 5 casos: son 116 casos de los que se pudieron
    comprobar 5.
    """
    r = procesar_estado_cuenta(requires_real_pdf("edocta-bbva"))
    regla = _regla(r.cobertura, "saldo_corrido")
    assert regla.aplicables == 116
    assert regla.evaluados == 5
    assert regla.motivo, "un hueco de 111 casos tiene que venir explicado"
    assert "116" in regla.resumen()


def test_el_renglon_que_siembra_la_cadena_cuenta_como_aplicable():
    """Ante la duda, el caso entra al denominador."""
    r = procesar_estado_cuenta(requires_real_pdf("edocta-bbva"))
    assert _regla(r.cobertura, "saldo_corrido").aplicables == 116  # no 115


def test_bajio_reporta_67_no_65():
    r = procesar_estado_cuenta(requires_real_pdf("edocta-bajio"))
    regla = _regla(r.cobertura, "saldo_corrido")
    assert (regla.aplicables, regla.evaluados) == (67, 65)


@pytest.mark.parametrize("nombre,aplicables", [
    ("edocta", 45), ("edocta-abril-santander", 110),
    ("edocta-julio-banorte", 283), ("edocta-inbursa", 44),
])
def test_los_que_ya_eran_honestos_no_cambian(nombre, aplicables):
    r = procesar_estado_cuenta(requires_real_pdf(nombre))
    regla = _regla(r.cobertura, "saldo_corrido")
    assert regla.aplicables == regla.evaluados == aplicables


# --- El universo de cada regla, medido fuera del codigo -----------------
def test_balanza_jerarquia_cuenta_los_padres_que_deberian_existir():
    """Dos cuentas de balanza.pdf declaran un padre que no esta en el
    documento: el par no se puede formar, pero el caso existe."""
    r = procesar_balanza(requires_real_pdf("balanza"))
    regla = _regla(r.cobertura, "jerarquia")
    assert regla.aplicables == 56   # 28 padres referidos x debe/haber
    assert regla.evaluados == 52    # 26 pares formados x 2


def test_balanza_businesspro_declara_la_partida_doble_que_no_corrio():
    r = procesar_balanza(requires_real_pdf("balanza-businesspro"))
    regla = _regla(r.cobertura, "partida_doble")
    assert regla.estado == NO_VERIFICABLE
    assert (regla.aplicables, regla.evaluados) == (1, 0)


def test_el_cruce_entre_documentos_declara_sus_49_cuentas():
    r = procesar_mayor(requires_real_pdf("mayor-gume"))
    regla = _regla(r.cobertura, "cruce_balanza")
    assert regla.estado == NO_VERIFICABLE
    assert (regla.aplicables, regla.evaluados) == (49, 0)


def test_el_total_declarado_que_el_documento_no_imprime():
    r = procesar_estado_cuenta(requires_real_pdf("edocta"))
    regla = _regla(r.cobertura, "total_declarado")
    assert regla.estado == NO_VERIFICABLE
    assert (regla.aplicables, regla.evaluados) == (2, 0)


@pytest.mark.lento          # 27 s
def test_polizas_incompletas_entran_al_denominador():
    """El PLAN dice que la cobertura las declara, y declarar exige estar
    en el denominador. En este fixture son 0, asi que apl == eval."""
    r = procesar_polizas(requires_real_pdf("poliza"))
    libro = r.libro
    incompletas = sum(1 for p in libro.polizas if not p.completa)
    regla = _regla(r.cobertura, "partida_doble")
    assert regla.aplicables == len(libro.polizas)
    assert regla.evaluados == len(libro.polizas) - incompletas


@pytest.mark.lento          # 16 s
def test_auxiliar_saldo_corrido_sobre_todos_los_movimientos():
    r = procesar_auxiliar(requires_real_pdf("auxiliar"))
    regla = _regla(r.cobertura, "saldo_corrido")
    assert regla.aplicables == regla.evaluados == 6783


@pytest.mark.lento
def test_auxiliar_gume_reporta_sobre_57024_no_sobre_21757():
    """El denominador es el documento entero, pase lo que pase arriba.

    En la 7f los evaluados eran 21,757 porque el pipeline no recalculaba;
    la 7g conecto `recalcular_saldos` y subieron a 47,987 con 26,032 saldos
    derivados de un ancla verificada. Lo que este test defiende es el
    DENOMINADOR: sigue siendo 57,024, y los 9,037 que faltan siguen
    contados y explicados.
    """
    r = procesar_auxiliar(requires_real_pdf("auxiliar-gume"))
    corrido = _regla(r.cobertura, "saldo_corrido")
    assert corrido.aplicables == 57024
    assert corrido.evaluados == 47987
    assert corrido.aplicables - corrido.evaluados == 9037
    assert corrido.motivo
    # 735 subtotales x debe/haber: las unidades de 'exactas'. En la
    # medicion previa este universo se anoto en subtotales (735) mientras
    # 'evaluados' iba en comprobaciones (344); mezclaba unidades.
    subtotales = _regla(r.cobertura, "subtotales")
    assert (subtotales.aplicables, subtotales.evaluados) == (1470, 344)


# --- Invariante transversal, sobre los cinco tipos ----------------------
# `lento` caso por caso y no en la funcion entera: la suite rapida sigue
# comprobando el denominador sobre estado de cuenta y mayor en cada ciclo,
# y los tres documentos caros van antes de entregar. Medido en la 8c:
# poliza 27 s por test, auxiliar 14 s, los tres de balanza 6 s juntos --
# contra menos de 0.5 s los otros dos casos. Fase 8c.
_TODOS = (
    pytest.param(procesar_balanza,
                 ("balanza", "balanza-businesspro", "balanza-gume"),
                 marks=pytest.mark.lento),
    pytest.param(procesar_auxiliar, ("auxiliar",), marks=pytest.mark.lento),
    pytest.param(procesar_polizas, ("poliza",), marks=pytest.mark.lento),
    (procesar_estado_cuenta, ("edocta", "edocta-bbva", "edocta-julio-banorte")),
    (procesar_mayor, ("mayor-gume",)),
)


@pytest.mark.parametrize("procesar,nombres", _TODOS)
def test_ninguna_regla_cuadra_sin_universo(procesar, nombres):
    for nombre in nombres:
        for regla in procesar(requires_real_pdf(nombre)).cobertura.reglas:
            if regla.estado == CUADRA:
                assert regla.aplicables is not None, (nombre, regla.regla)
            if regla.aplicables is not None:
                assert regla.aplicables >= regla.evaluados, (nombre, regla.regla)
            assert regla.exactas + len(regla.con_tolerancia) \
                + len(regla.discrepancias) == regla.evaluados, (nombre, regla.regla)


@pytest.mark.parametrize("procesar,nombres", _TODOS)
def test_ningun_conteo_se_imprime_sin_su_denominador(procesar, nombres):
    for nombre in nombres:
        cobertura = procesar(requires_real_pdf(nombre)).cobertura
        for regla in cobertura.reglas:
            texto = regla.resumen()
            if regla.aplicables is None:
                assert "sin determinar" in texto, (nombre, regla.regla)
            else:
                assert f"de {regla.aplicables}" in texto, (nombre, regla.regla)
        total = sum(r.aplicables or 0 for r in cobertura.reglas)
        assert f"de {total}" in cobertura.resumen(), nombre


@pytest.mark.parametrize("procesar,nombres", _TODOS)
def test_un_hueco_entre_aplicables_y_evaluados_viene_explicado(procesar, nombres):
    """Si la regla no corrio en todo el universo, hay que decir por que."""
    for nombre in nombres:
        for regla in procesar(requires_real_pdf(nombre)).cobertura.reglas:
            if regla.aplicables is None:
                continue
            if regla.aplicables > regla.evaluados:
                assert regla.motivo, (nombre, regla.regla)


# --- La hoja Validacion de los cinco exportadores -----------------------
@pytest.mark.lento          # 55 s
def test_los_cinco_exportadores_llevan_la_columna(tmp_path):
    """Sin denominador en la hoja, el Excel repite la misma mentira."""
    import openpyxl

    from contapdf.export.excel import (
        exportar_auxiliar, exportar_balanza, exportar_estado_cuenta,
        exportar_mayor, exportar_polizas,
    )
    casos = (
        ("balanza", procesar_balanza, "balanza", exportar_balanza),
        ("auxiliar", procesar_auxiliar, "auxiliar", exportar_auxiliar),
        ("poliza", procesar_polizas, "libro", exportar_polizas),
        ("edocta-bbva", procesar_estado_cuenta, "estado", exportar_estado_cuenta),
        ("mayor-gume", procesar_mayor, "mayor", exportar_mayor),
    )
    for nombre, procesar, campo, exportar in casos:
        r = procesar(requires_real_pdf(nombre))
        destino = tmp_path / f"{nombre}.xlsx"
        exportar(getattr(r, campo), r.cobertura, destino)
        hoja = openpyxl.load_workbook(destino)["Validacion"]
        columnas = [c.value for c in hoja[1]]
        assert columnas[:5] == ["regla", "estado", "aplicables", "evaluados",
                                "exactos"], nombre
        filas = {f[0]: f for f in hoja.iter_rows(min_row=2, max_col=5,
                                                 values_only=True) if f[0]}
        for regla in r.cobertura.reglas:
            assert filas[regla.regla][2] == regla.aplicables, (nombre, regla.regla)
            assert filas[regla.regla][3] == regla.evaluados, (nombre, regla.regla)


def test_el_excel_de_bbva_dice_5_de_116(tmp_path):
    import openpyxl

    from contapdf.export.excel import exportar_estado_cuenta

    r = procesar_estado_cuenta(requires_real_pdf("edocta-bbva"))
    destino = tmp_path / "bbva.xlsx"
    exportar_estado_cuenta(r.estado, r.cobertura, destino)
    hoja = openpyxl.load_workbook(destino)["Validacion"]
    fila = next(f for f in hoja.iter_rows(min_row=2, values_only=True)
                if f[0] == "saldo_corrido")
    assert (fila[2], fila[3]) == (116, 5)

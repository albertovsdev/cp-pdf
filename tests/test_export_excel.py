"""Exportacion a Excel."""

from __future__ import annotations

from decimal import Decimal

import openpyxl
from conftest import synthetic_document

from contapdf.export.excel import exportar_balanza
from contapdf.parsers.balanza import BalanzaParser
from contapdf.validate.rules import evaluar_balanza


def _exportar(tmp_path, name="balanza_sintetica"):
    balanza = BalanzaParser().parse(synthetic_document(name))
    cobertura = evaluar_balanza(balanza)
    destino = tmp_path / "salida.xlsx"
    return (balanza, list(cobertura.discrepancias),
            exportar_balanza(balanza, cobertura, destino))


def test_genera_el_archivo_en_la_ruta_pedida(tmp_path):
    _, _, destino = _exportar(tmp_path)
    assert destino == tmp_path / "salida.xlsx"
    assert destino.exists()


def test_la_hoja_se_llama_balanza_y_trae_encabezados(tmp_path):
    _, _, destino = _exportar(tmp_path)
    wb = openpyxl.load_workbook(destino)
    assert wb.sheetnames[0] == "Balanza"
    ws = wb["Balanza"]
    assert [c.value for c in ws[1]] == [
        "cuenta", "nivel", "cuenta_padre", "naturaleza", "nombre",
        "saldo_ini_deudor", "saldo_ini_acreedor", "debe", "haber",
        "saldo_fin_deudor", "saldo_fin_acreedor", "es_acumulativa",
    ]


def test_marca_las_filas_acumulativas(tmp_path):
    # El contador filtra la hoja plana: sin esta columna no puede saber
    # cuales son subtotales y sumaria dos veces.
    balanza, _, destino = _exportar(tmp_path)
    ws = openpyxl.load_workbook(destino)["Balanza"]
    columna = [c.value for c in ws["L"][1:]]
    assert columna == [f.es_acumulativa for f in balanza.filas]
    assert any(columna)


def test_el_encabezado_queda_congelado(tmp_path):
    _, _, destino = _exportar(tmp_path)
    assert openpyxl.load_workbook(destino)["Balanza"].freeze_panes == "A2"


def test_los_montos_se_reabren_como_numeros_no_como_texto(tmp_path):
    balanza, _, destino = _exportar(tmp_path)
    ws = openpyxl.load_workbook(destino)["Balanza"]
    for fila, row in zip(balanza.filas, ws.iter_rows(min_row=2, max_row=1 + len(balanza.filas))):
        for celda, esperado in zip(row[5:11],
                                   (fila.saldo_ini_deudor, fila.saldo_ini_acreedor,
                                    fila.debe, fila.haber,
                                    fila.saldo_fin_deudor, fila.saldo_fin_acreedor)):
            assert isinstance(celda.value, (int, float)), f"{celda.coordinate} es texto"
            assert abs(Decimal(str(celda.value)) - esperado) < Decimal("0.005")
            assert celda.number_format == "#,##0.00"


def test_conserva_negativos(tmp_path):
    _, _, destino = _exportar(tmp_path)
    ws = openpyxl.load_workbook(destino)["Balanza"]
    negativos = [c.value for row in ws.iter_rows(min_row=2) for c in row
                 if isinstance(c.value, (int, float)) and c.value < 0]
    assert negativos == [-1250.25]


def test_sin_discrepancias_la_hoja_no_lista_ninguna(tmp_path):
    _, discrepancias, destino = _exportar(tmp_path)
    assert discrepancias == []
    ws = openpyxl.load_workbook(destino)["Validacion"]
    filas = [[c.value for c in f] for f in ws.iter_rows()]
    encabezado = next(i for i, f in enumerate(filas) if f[0] == "fila")
    assert filas[encabezado + 1:] == []


def test_con_discrepancias_las_lista_bajo_la_cobertura(tmp_path):
    _, discrepancias, destino = _exportar(tmp_path, "balanza_descuadrada")
    assert len(discrepancias) == 1
    ws = openpyxl.load_workbook(destino)["Validacion"]
    filas = [[c.value for c in f] for f in ws.iter_rows()]
    # Desde la fase 7f la hoja lleva el denominador: sin 'aplicables' una
    # regla que corrio en el 4% del documento se lee igual que una entera.
    assert filas[0][:5] == ["regla", "estado", "aplicables", "evaluados",
                            "exactos"]
    encabezado = next(i for i, f in enumerate(filas) if f[0] == "fila")
    assert filas[encabezado][:4] == ["fila", "regla", "esperado", "obtenido"]
    assert filas[encabezado + 1][:2] == ["102-02", "renglon"]


def test_marca_la_fila_afectada(tmp_path):
    balanza, _, destino = _exportar(tmp_path, "balanza_descuadrada")
    ws = openpyxl.load_workbook(destino)["Balanza"]
    fila_mala = next(i for i, f in enumerate(balanza.filas) if f.cuenta == "102-02")

    marcada = ws.cell(row=2 + fila_mala, column=1)
    limpia = ws.cell(row=2, column=1)
    assert marcada.fill.fgColor.rgb != limpia.fill.fgColor.rgb


def test_no_imprime(tmp_path, capsys):
    _exportar(tmp_path)
    salida = capsys.readouterr()
    assert salida.out == "" and salida.err == ""


def test_la_hoja_validacion_existe_aunque_no_haya_discrepancias(tmp_path):
    # PLAN 2: nunca reportar un resultado sin su cobertura.
    _, discrepancias, destino = _exportar(tmp_path)
    assert discrepancias == []
    wb = openpyxl.load_workbook(destino)
    assert "Validacion" in wb.sheetnames
    filas = [[c.value for c in fila] for fila in wb["Validacion"].iter_rows()]
    texto = " ".join(str(v) for f in filas for v in f if v is not None)
    assert "renglon" in texto and "cuadra" in texto


def test_la_hoja_validacion_lista_las_cuatro_reglas_y_las_discrepancias(tmp_path):
    _, discrepancias, destino = _exportar(tmp_path, "balanza_descuadrada")
    ws = openpyxl.load_workbook(destino)["Validacion"]
    texto = " ".join(str(c.value) for fila in ws.iter_rows() for c in fila
                     if c.value is not None)
    for regla in ("renglon", "jerarquia", "totales", "partida_doble"):
        assert regla in texto
    assert "102-02" in texto


def test_la_naturaleza_sin_determinar_sale_vacia_en_el_excel(tmp_path):
    from contapdf.ir import Word
    from contapdf.parsers.balanza import BalanzaParser
    from contapdf.validate.rules import evaluar_balanza
    import dataclasses

    balanza = BalanzaParser().parse(synthetic_document("balanza_sintetica"))
    filas = list(balanza.filas)
    filas[0] = dataclasses.replace(filas[0], naturaleza="",
                                   naturaleza_origen="sin_determinar")
    balanza = dataclasses.replace(balanza, filas=tuple(filas))
    destino = tmp_path / "n.xlsx"
    exportar_balanza(balanza, evaluar_balanza(balanza), destino)

    ws = openpyxl.load_workbook(destino)["Balanza"]
    assert ws.cell(row=2, column=4).value in (None, "")


def test_el_excel_no_exporta_la_procedencia(tmp_path):
    # Duplica el ancho de la hoja y el contador la ignora.
    _, _, destino = _exportar(tmp_path)
    encabezados = [c.value for c in openpyxl.load_workbook(destino)["Balanza"][1]]
    assert "naturaleza" in encabezados
    assert not any("origen" in str(h) for h in encabezados)


def test_el_libro_diario_sale_en_cuatro_hojas(tmp_path):
    from conftest import requires_real_pdf

    from contapdf.export.excel import exportar_polizas
    from contapdf.extract.strategy import extraer
    from contapdf.parsers.polizas import PolizasParser
    from contapdf.validate.rules import evaluar_polizas

    doc, _ = extraer(requires_real_pdf("poliza"), page_numbers=[1, 2])
    libro = PolizasParser().parse(doc)
    destino = tmp_path / "diario.xlsx"
    exportar_polizas(libro, evaluar_polizas(libro), destino)

    wb = openpyxl.load_workbook(destino)
    assert wb.sheetnames == ["Polizas", "Movimientos", "CFDI", "Plana",
                             "Validacion"]
    plana = wb["Plana"]
    # La hoja plana repite el encabezado en cada movimiento: es la que el
    # contador filtra.
    assert plana.max_row == len(libro.movimientos) + 1
    encabezados = [c.value for c in plana[1]]
    assert "poliza_id" in encabezados and "cuenta" in encabezados
    assert "total_debe" in encabezados


def test_las_hojas_exportan_los_campos_con_los_que_se_identifica_la_poliza(tmp_path):
    from conftest import requires_real_pdf

    from contapdf.export.excel import exportar_polizas
    from contapdf.extract.strategy import extraer
    from contapdf.parsers.polizas import PolizasParser
    from contapdf.validate.rules import evaluar_polizas

    doc, _ = extraer(requires_real_pdf("poliza"), page_numbers=[1, 2])
    libro = PolizasParser().parse(doc)
    destino = tmp_path / "id.xlsx"
    exportar_polizas(libro, evaluar_polizas(libro), destino)

    wb = openpyxl.load_workbook(destino)
    for hoja in ("Polizas", "Plana"):
        encabezados = [c.value for c in wb[hoja][1]]
        assert {"tipo", "fecha", "descripcion", "folio"} <= set(encabezados)
    # Y el valor llega, no solo la columna.
    fila = [c.value for c in wb["Polizas"][2]]
    assert "18243" in [str(v) for v in fila]


def test_el_libro_mayor_sale_en_dos_hojas_mas_la_plana(tmp_path):
    from conftest import requires_real_pdf

    from contapdf.export.excel import exportar_mayor
    from contapdf.extract.strategy import extraer
    from contapdf.parsers.mayor import MayorParser
    from contapdf.validate.rules import evaluar_mayor

    doc, _ = extraer(requires_real_pdf("mayor-gume"))
    mayor = MayorParser().parse(doc)
    destino = tmp_path / "mayor.xlsx"
    exportar_mayor(mayor, evaluar_mayor(mayor), destino)

    wb = openpyxl.load_workbook(destino)
    assert wb.sheetnames == ["Cuentas", "Meses", "Plana", "Validacion"]
    assert wb["Cuentas"].max_row == len(mayor.cuentas) + 1
    assert wb["Meses"].max_row == len(mayor.meses) + 1
    assert wb["Plana"].max_row == len(mayor.meses) + 1
    encabezados = [c.value for c in wb["Plana"][1]]
    assert "saldo_inicial" in encabezados and "periodo" in encabezados

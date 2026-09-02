"""CLI: correr el pipeline sobre un PDF y reportar en la terminal."""

from __future__ import annotations

import io
from decimal import Decimal

import openpyxl
from conftest import requires_real_pdf

from contapdf.cli import main, reportar
from contapdf.parsers.balanza import Balanza, FilaBalanza, Totales
from contapdf.validate.rules import (
    CUADRA,
    FALLA,
    Cobertura,
    Discrepancia,
    ResultadoRegla,
)


def _fila(cuenta: str) -> FilaBalanza:
    cero = Decimal("0.00")
    return FilaBalanza(cuenta=cuenta, nivel=1, cuenta_padre="", naturaleza="D",
                       nombre="Caja", saldo_ini_deudor=cero,
                       saldo_ini_acreedor=cero, debe=Decimal("1234.56"),
                       haber=cero, saldo_fin_deudor=Decimal("1234.56"),
                       saldo_fin_acreedor=cero)


def _cobertura(discrepancias=()):
    reglas = tuple(
        ResultadoRegla(regla=nombre, estado=FALLA if discrepancias and nombre == "renglon"
                       else CUADRA, aplicables=2, evaluados=2, exactas=2,
                       discrepancias=tuple(discrepancias) if nombre == "renglon" else ())
        for nombre in ("renglon", "jerarquia", "totales", "partida_doble"))
    return Cobertura(reglas=reglas)


def _reporte(discrepancias=(), totales=Totales(Decimal("1234.56"), Decimal("1234.56")),
             mapeo=None):
    salida = io.StringIO()
    balanza = Balanza(filas=(_fila("101"), _fila("102")), totales=totales, mapeo=mapeo)
    reportar("balanza.pdf", 9, "pdf_text", balanza, _cobertura(discrepancias),
             None, salida)
    return salida.getvalue()


def test_reporta_paginas_filas_y_totales():
    texto = _reporte()
    assert "9" in texto
    assert "2" in texto
    assert "1,234.56" in texto


def test_avisa_cuando_el_pdf_no_trae_fila_de_totales():
    assert "sin fila de totales" in _reporte(totales=None).lower()


def test_sin_discrepancias_lo_dice():
    assert "sin discrepancias" in _reporte().lower()


def test_lista_las_discrepancias_con_esperado_y_obtenido():
    texto = _reporte([Discrepancia(fila="102-02", indice=7, regla="renglon",
                                   esperado=Decimal("40749.50"),
                                   obtenido=Decimal("40849.50"))])
    assert "102-02" in texto
    assert "renglon" in texto
    assert "40,749.50" in texto and "40,849.50" in texto


def test_main_sobre_el_pdf_real(tmp_path):
    pdf = requires_real_pdf("balanza")
    destino = tmp_path / "balanza.xlsx"
    salida = io.StringIO()

    codigo = main(["balanza", str(pdf), "-o", str(destino)], salida=salida)

    assert codigo == 0                     # 0 = cuadra
    assert destino.exists()
    assert openpyxl.load_workbook(destino).sheetnames == ["Balanza", "Validacion"]
    texto = salida.getvalue()
    assert "475" in texto                  # filas del documento real
    assert "sin discrepancias" in texto.lower()
    assert str(destino) in texto


def test_main_sin_o_no_escribe_excel(tmp_path):
    pdf = requires_real_pdf("balanza")
    salida = io.StringIO()
    assert main(["balanza", str(pdf)], salida=salida) == 0
    assert list(tmp_path.iterdir()) == []


def test_main_avisa_si_el_pdf_no_existe(tmp_path):
    salida = io.StringIO()
    codigo = main(["balanza", str(tmp_path / "no-existe.pdf")], salida=salida)
    assert codigo == 2
    assert "no existe" in salida.getvalue().lower()


def test_el_codigo_de_salida_distingue_documento_descuadrado():
    # 0 cuadra, 1 hay discrepancias, 2 no se pudo procesar: para poder
    # encadenarlo en un script sin leer el texto.
    from contapdf.cli import codigo_de_salida

    assert codigo_de_salida(_cobertura()) == 0
    assert codigo_de_salida(_cobertura(
        [Discrepancia("101", 0, "renglon", Decimal(1), Decimal(2))])) == 1


def test_main_sobre_business_pro_elige_la_extraccion_por_corridas(tmp_path):
    pdf = requires_real_pdf("balanza-businesspro")
    destino = tmp_path / "bp.xlsx"
    salida = io.StringIO()

    codigo = main(["balanza", str(pdf), "-o", str(destino)], salida=salida)

    texto = salida.getvalue()
    assert codigo == 0
    assert "225" in texto
    assert "pdf_chars" in texto
    assert "sin discrepancias" in texto.lower()
    assert openpyxl.load_workbook(destino).sheetnames == ["Balanza", "Validacion"]


def test_reporta_la_estrategia_de_extraccion():
    texto = _reporte()
    assert "extraccion" in texto.lower()


def test_el_reporte_nunca_dice_cero_discrepancias_a_secas():
    texto = _reporte()
    assert "0 discrepancias" not in texto
    assert "cobertura" in texto.lower()
    assert "4 reglas" in texto


def test_el_reporte_marca_lo_que_solo_se_apoya_en_vocabulario():
    from contapdf.parsers.balanza import Mapeo

    salida = io.StringIO()
    balanza = Balanza(filas=(_fila("101"),), totales=None,
                      mapeo=Mapeo(campos={}, forma="saldo_con_signo",
                                  verificado_por="vocabulario",
                                  orientacion_verificada=False,
                                  filas_afectadas=725))
    reportar("x.pdf", 1, "pdf_text", balanza, _cobertura(), None, salida)
    texto = salida.getvalue().lower()
    assert "vocabulario" in texto
    assert "725" in texto


def test_main_sobre_gume_reporta_cobertura_completa(tmp_path):
    pdf = requires_real_pdf("balanza-gume")
    salida = io.StringIO()
    codigo = main(["balanza", str(pdf)], salida=salida)
    texto = salida.getvalue()
    assert codigo == 0
    assert "734" in texto
    assert "4 reglas" in texto
    assert "vocabulario" in texto.lower()


def test_el_reporte_incluye_la_procedencia_de_la_naturaleza():
    texto = _reporte()
    assert "naturaleza" in texto.lower()


def test_main_con_plantillas_reporta_la_plantilla(tmp_path):
    pdf = requires_real_pdf("balanza-gume")
    salida = io.StringIO()
    codigo = main(["balanza", str(pdf), "--tenant", "despacho-a",
                   "--plantillas", str(tmp_path)], salida=salida)
    texto = salida.getvalue().lower()
    assert codigo == 0
    assert "plantilla" in texto
    assert "pendiente" in texto

    segunda = io.StringIO()
    main(["balanza", str(pdf), "--tenant", "despacho-a",
          "--plantillas", str(tmp_path)], salida=segunda)
    assert "reutiliz" in segunda.getvalue().lower()


def test_confirmar_desde_el_cli(tmp_path):
    pdf = requires_real_pdf("balanza-gume")
    main(["balanza", str(pdf), "--tenant", "t", "--plantillas", str(tmp_path)],
         salida=io.StringIO())

    from contapdf.templates.store import AlmacenPlantillas

    huella = AlmacenPlantillas(tmp_path).listar("t")[0].huella
    salida = io.StringIO()
    codigo = main(["confirmar", "--tenant", "t", "--plantillas", str(tmp_path),
                   "--huella", huella, "--por", "contadora"], salida=salida)
    assert codigo == 0
    assert AlmacenPlantillas(tmp_path).buscar("t", huella).confirmada_por == "contadora"

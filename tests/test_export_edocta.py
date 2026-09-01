"""El estado de cuenta a Excel: cuentas y movimientos, relacionados.

Fase 7e. Era el unico de los cinco parsers sin salida a Excel.

Cuatro hojas, la misma forma que polizas y mayor: las relacionadas por un
lado y una plana denormalizada por otro, que es la que el contador filtra.
La metadata del documento (banco, RFC, periodo) va como columnas de la hoja
Cuentas: son de una a tres filas, asi que repetirla no cuesta nada y evita
una quinta hoja que nadie abriria.
"""

from __future__ import annotations

from decimal import Decimal

import openpyxl
import pytest
from conftest import requires_real_pdf

from contapdf.export.excel import exportar_estado_cuenta
from contapdf.extract.strategy import extraer
from contapdf.parsers.estado_cuenta import EstadoCuentaParser
from contapdf.validate.rules import evaluar_estado_cuenta

_CON_TABLA = ("edocta", "edocta-abril-santander", "edocta-julio-banorte",
              "edocta-bajio", "edocta-inbursa", "edocta-bbva")


def _exportado(nombre, tmp_path):
    doc, _ = extraer(requires_real_pdf(nombre))
    estado = EstadoCuentaParser().parse(doc)
    destino = tmp_path / f"{nombre}.xlsx"
    exportar_estado_cuenta(estado, evaluar_estado_cuenta(estado), destino)
    return estado, openpyxl.load_workbook(destino)


# --- Criterio 1 ---------------------------------------------------------
@pytest.mark.parametrize("nombre", _CON_TABLA)
def test_los_seis_salen_a_excel(nombre, tmp_path):
    estado, libro = _exportado(nombre, tmp_path)
    assert libro.sheetnames == ["Cuentas", "Movimientos", "Plana", "Validacion"]
    assert libro["Cuentas"].max_row == len(estado.cuentas) + 1
    assert libro["Movimientos"].max_row == len(estado.movimientos) + 1
    assert libro["Plana"].max_row == len(estado.movimientos) + 1


def test_las_dos_tablas_quedan_relacionadas(tmp_path):
    """Ninguna fila huerfana: todo movimiento apunta a una cuenta que existe."""
    estado, libro = _exportado("edocta-julio-banorte", tmp_path)
    cuentas = libro["Cuentas"]
    columnas = [c.value for c in cuentas[1]]
    cual = columnas.index("num_cuenta")
    declaradas = {fila[cual].value for fila in cuentas.iter_rows(min_row=2)}
    assert len(declaradas) == 2

    movimientos = libro["Movimientos"]
    donde = [c.value for c in movimientos[1]].index("num_cuenta")
    usadas = {fila[donde].value for fila in movimientos.iter_rows(min_row=2)}
    assert usadas <= declaradas


def test_la_plana_repite_la_cuenta_en_cada_movimiento(tmp_path):
    estado, libro = _exportado("edocta-julio-banorte", tmp_path)
    plana = libro["Plana"]
    columnas = [c.value for c in plana[1]]
    for campo in ("num_cuenta", "producto", "saldo_inicial", "descripcion",
                  "deposito", "retiro", "saldo"):
        assert campo in columnas, campo
    # El encabezado de la cuenta viaja en cada renglon: es lo que permite
    # filtrar por cuenta sin abrir otra hoja.
    fila = next(plana.iter_rows(min_row=2, values_only=True))
    assert fila[columnas.index("producto")] == "ENLACE NEGOCIOS PFAE"


def test_la_hoja_cuentas_trae_la_metadata_del_documento(tmp_path):
    estado, libro = _exportado("edocta", tmp_path)
    columnas = [c.value for c in libro["Cuentas"][1]]
    for campo in ("banco", "rfc", "periodo_ini", "periodo_fin"):
        assert campo in columnas, campo
    fila = next(libro["Cuentas"].iter_rows(min_row=2, values_only=True))
    assert fila[columnas.index("periodo_ini")] == "01 ABR 2025"


def test_el_dinero_sale_como_numero_con_formato(tmp_path):
    """Numero con formato de moneda, nunca texto: el contador va a sumarlo.

    openpyxl guarda el Decimal como numero de celda y lo devuelve como
    float al releer; lo que se comprueba es que el VALOR sobreviva exacto y
    que la celda lleve el formato, igual que en balanza y polizas.
    """
    estado, libro = _exportado("edocta", tmp_path)
    hoja = libro["Movimientos"]
    donde = [c.value for c in hoja[1]].index("saldo")
    celda = hoja.cell(row=2, column=donde + 1)
    assert isinstance(celda.value, (int, float))
    assert Decimal(str(celda.value)) == estado.movimientos[0].saldo
    assert celda.number_format == "#,##0.00"


def test_la_validacion_lleva_las_cuatro_reglas(tmp_path):
    estado, libro = _exportado("edocta", tmp_path)
    reglas = {fila[0] for fila in
              libro["Validacion"].iter_rows(min_row=2, max_col=1, values_only=True)}
    assert {"resumen", "resumen_movimientos", "saldo_corrido",
            "total_declarado"} <= reglas


def test_lo_no_verificable_lleva_su_motivo(tmp_path):
    estado, libro = _exportado("edocta", tmp_path)
    for regla, estado_regla, detalle in libro["Validacion"].iter_rows(
            min_row=2, max_col=3, values_only=True):
        if estado_regla == "no_verificable":
            assert detalle, regla


def test_un_saldo_ilegible_sale_vacio_y_no_como_cero(tmp_path):
    """BBVA imprime el saldo corrido solo una vez por dia."""
    estado, libro = _exportado("edocta-bbva", tmp_path)
    hoja = libro["Movimientos"]
    donde = [c.value for c in hoja[1]].index("saldo")
    vacios = sum(1 for fila in hoja.iter_rows(min_row=2, values_only=True)
                 if fila[donde] is None)
    assert vacios == sum(1 for m in estado.movimientos if m.saldo is None) > 0

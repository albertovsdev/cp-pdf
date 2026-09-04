"""El exportador no puede ser cuadratico en el numero de renglones.

Fase 8c. Medido: `auxiliar-gume` tardaba **1 243 s en escribir el .xlsx**
contra 182 s en leerlo y validarlo. La causa era `hoja[hoja.max_row]`
dentro del bucle por renglon, que recorre la hoja entera DOS veces por
renglon: una en `max_row` y otra en el `max_column` que `hoja[n]` usa por
dentro. Un bucle lineal sobre dos llamadas lineales da un coste
cuadratico. Medido aparte: 1 000 llamadas a `max_row` cuestan 0.040 s con
1 000 renglones y 0.300 s con 8 000.

Estaba en tres bucles y afectaba a **los cinco exportadores**: `_hoja` lo
usan polizas, mayor, estado de cuenta y auxiliar; `exportar_balanza` tiene
su propio bucle; y la hoja `Validacion` la escriben los cinco.

Dos guardas, porque protegen cosas distintas:

- La **estructural** es exacta y cuesta milisegundos: cuenta los accesos a
  `max_row` y `max_column` y comprueba que no crecen con los renglones. Va
  en la suite rapida.
- La **de escalado** mide tiempo de verdad y atrapa un cuadratico que
  llegue por otro camino. Es lenta por definicion —hay que escribir
  bastantes renglones para que el termino cuadratico se note— asi que va
  marcada `lento`.

Sin la segunda, un cuadratico nuevo que no pase por esas dos propiedades
no se veria.
Sin la primera, la unica red seria una medicion de tiempo, que es
justamente lo que no se debe usar como test en cada ciclo.
"""

from __future__ import annotations

import time
from decimal import Decimal

import pytest
from openpyxl.worksheet.worksheet import Worksheet

from contapdf.export.excel import exportar_auxiliar
from contapdf.parsers.auxiliar import Auxiliar, FilaAuxiliar
from contapdf.validate.rules import Cobertura


def _auxiliar(renglones: int) -> Auxiliar:
    """Datos sinteticos: esto mide el exportador, no el parser."""
    filas = tuple(
        FilaAuxiliar(
            cuenta=f"1120-001-{i % 900:03d}",
            nombre_cuenta=f"CUENTA DE PRUEBA {i}",
            saldo_inicial_cuenta=Decimal("100.00"),
            folio=str(i), fecha="2025-01-15", tipo_movimiento="Dr",
            documento=f"F-{i}", tercero="TERCERO SA DE CV",
            concepto="concepto de prueba suficientemente largo",
            debe=Decimal("1234.56"), haber=None,
            saldo=Decimal("9876.54"), es_subtotal=False,
            pagina=1 + i // 60, top=float(i % 60),
        )
        for i in range(renglones)
    )
    return Auxiliar(filas=filas, secciones=(), forma="", mapeo=None)


@pytest.fixture()
def cobertura() -> Cobertura:
    return Cobertura(reglas=())


#: Las dos propiedades que recorren la hoja entera. `max_row` era la
#: primera causa; `max_column` es la segunda, y llega sola: `hoja[n]`
#: pasa por `iter_rows(..., max_col=self.max_column)`, asi que cambiar
#: `max_row` por `_current_row` deja el cuadratico donde estaba. Medido:
#: con ese arreglo a medias, `auxiliar-gume` bajo de 1 197 s a 587 s; con
#: el bucle sin indexar la hoja, `auxiliar` bajo de 14.11 s a 0.63 s.
_LINEALES = ("max_row", "max_column")


class _Contador:
    """Cuenta los recorridos de hoja entera sin cambiar lo que devuelven."""

    def __init__(self) -> None:
        self.accesos = 0
        self._originales = {n: getattr(Worksheet, n) for n in _LINEALES}

    def __enter__(self) -> "_Contador":
        contador = self

        for nombre, original in self._originales.items():
            def espia(hoja, _original=original):
                contador.accesos += 1
                return _original.fget(hoja)

            setattr(Worksheet, nombre, property(espia))
        return self

    def __exit__(self, *_) -> None:
        for nombre, original in self._originales.items():
            setattr(Worksheet, nombre, original)


def _accesos_para(renglones: int, cobertura: Cobertura, destino) -> int:
    with _Contador() as contador:
        exportar_auxiliar(_auxiliar(renglones), cobertura,
                          destino / f"{renglones}.xlsx")
    return contador.accesos


def test_no_se_recorre_la_hoja_entera_una_vez_por_renglon(tmp_path, cobertura):
    """La guarda exacta: los accesos no pueden crecer con los renglones.

    `max_row` y `max_column` recorren la hoja entera. Llamar a cualquiera
    de los dos dentro del bucle es lo que convirtio 182 s de lectura en
    1 243 s de escritura.
    """
    pocos = _accesos_para(200, cobertura, tmp_path)
    muchos = _accesos_para(2000, cobertura, tmp_path)

    assert muchos == pocos, (
        f"con 200 renglones se recorre la hoja entera {pocos} veces y con "
        f"2 000, {muchos}: hay un max_row o un max_column dentro del bucle "
        f"y el exportador vuelve a ser cuadratico")


def test_el_coste_por_renglon_no_crece(tmp_path, cobertura):
    """Barato y sin reloj: el trabajo por renglon tiene que ser constante.

    Se cuentan accesos, no segundos, para que valga en una maquina cargada.
    """
    accesos = [_accesos_para(n, cobertura, tmp_path) for n in (100, 400, 1600)]
    assert len(set(accesos)) == 1, (
        f"los recorridos de hoja entera cambian con el tamano: {accesos}")


@pytest.mark.lento
def test_escribir_el_doble_de_renglones_no_cuesta_el_cuadruple(tmp_path,
                                                               cobertura):
    """La guarda de escalado, por si el cuadratico vuelve por otra puerta.

    Medido con el defecto: duplicar renglones multiplicaba el tiempo por
    3.5 (1 000 -> 0.35 s, 2 000 -> 1.22 s, 4 000 -> 4.28 s). Lineal daria
    2.0. El umbral esta en 2.6, a mitad del hueco entre lo lineal y lo
    medido con el defecto, para que no falle por ruido de la maquina.
    """
    def segundos(renglones: int) -> float:
        datos = _auxiliar(renglones)
        # Dos pasadas y la mejor: en una maquina compartida una sola
        # medicion de tiempo no es una medicion.
        return min(_una_pasada(datos, cobertura, tmp_path, renglones, i)
                   for i in range(2))

    base = segundos(3000)
    doble = segundos(6000)
    factor = doble / base
    assert factor < 2.6, (
        f"duplicar los renglones costo {factor:.2f}x ({base:.2f}s -> "
        f"{doble:.2f}s): el exportador volvio a escalar peor que lineal")


def _una_pasada(datos, cobertura, destino, renglones: int, intento: int) -> float:
    comenzo = time.perf_counter()
    exportar_auxiliar(datos, cobertura, destino / f"t{renglones}-{intento}.xlsx")
    return time.perf_counter() - comenzo
